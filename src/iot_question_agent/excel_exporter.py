from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


DEFAULT_COLUMNS = [
    "Вопрос",
    "Правильный ответ",
    "Неправильный ответ 1",
    "Неправильный ответ 2",
    "Раздел",
    "Пункт",
    "Пояснение правильного ответа",
    "Тема",
]


def export_questions_xlsx(
    path: Path,
    questions: list[dict[str, str]],
    columns: Iterable[str] = DEFAULT_COLUMNS,
) -> None:
    columns = list(columns)
    wb = Workbook()
    ws = wb.active
    ws.title = "Вопросы"

    ws.append(columns)
    for q in questions:
        ws.append([q.get(col, "") for col in columns])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    thin = Side(style="thin", color="D0D7DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    widths = {
        "Вопрос": 52,
        "Правильный ответ": 42,
        "Неправильный ответ 1": 42,
        "Неправильный ответ 2": 42,
        "Раздел": 34,
        "Пункт": 12,
        "Пояснение правильного ответа": 58,
        "Тема": 26,
    }

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    for idx, col_name in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(col_name, 24)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 60
    ws.row_dimensions[1].height = 28

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
